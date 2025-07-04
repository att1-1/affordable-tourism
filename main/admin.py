# pylint: disable=no-member
from django.contrib import admin
from django.urls import path
from django.shortcuts import render, redirect
from django import forms
from django.http import HttpResponse
from django.db import transaction
from openpyxl import Workbook
from pandas import read_excel
from .models import Route, AgeGroup, Season, Skill, Comment
from io import BytesIO
from openpyxl.utils import get_column_letter


class ExportImportMixin:
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('export-excel/', self.export_excel),
            path('import-excel/', self.import_excel),
        ]
        return custom_urls + urls

    def export_excel(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Маршруты"

        columns = [
            'ID', 'Название', 'Возрастные ступени', 'Протяженность (км)', 
            'Продолжительность (часы)', 'Способ передвижения', 'Стоимость (руб)',
            'Даты проведения', 'Сезоны', 'Навыки', 'Организатор', 'Телефон', 
            'Почта', 'Организация', 'Паспорт', 'Карта', 'Изображение'
        ]
        ws.append(columns)

        for route in Route.objects.all():
            ws.append([
                route.id,
                route.name,
                ','.join(str(ag.code) for ag in route.age_groups.all()),
                route.distance,
                route.duration,
                route.transportation_method,
                route.cost,
                route.event_dates,
                ','.join(s.name for s in route.seasons.all()),
                ','.join(sk.code for sk in route.skills.all()),
                route.organizer_name,
                route.organizer_phone,
                route.organizer_email,
                route.organization,
                route.route_passport,
                route.map_link,
                route.image.url if route.image else ''
            ])

        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)

            for cell in col:
                try:
                    value = str(cell.value) if cell.value else ""
                    max_length = max(max_length, len(value))
                except:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        for col in [1, 4, 5, 7]:
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 12

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=routes_export.xlsx'
        return response

    def import_excel(self, request):
        class ImportForm(forms.Form):
            file = forms.FileField(label='Excel файл')

        if request.method == 'POST':
            form = ImportForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    df = read_excel(request.FILES['file'])
                    with transaction.atomic():
                        for _, row in df.iterrows():
                            self.process_row(row)
                        self.message_user(request, "Данные успешно обновлены!")
                except Exception as e:
                    self.message_user(request, f"Ошибка: {str(e)}", level='ERROR')
                return redirect('..')
        else:
            form = ImportForm()

        return render(request, 'admin/import_excel.html', {'form': form})

    def process_row(self, row):
        route, created = Route.objects.update_or_create(
            id=row['ID'],
            defaults={
                'name': row['Название'],
                'distance': row['Протяженность (км)'],
                'duration': row['Продолжительность (часы)'],
                'transportation_method': row['Способ передвижения'],
                'cost': row['Стоимость (руб)'],
                'event_dates': row['Даты проведения'],
                'organizer_name': row['Организатор'],
                'organizer_phone': row['Телефон'],
                'organizer_email': row['Почта'],
                'organization': row['Организация'],
                'route_passport': row['Паспорт'],
                'map_link': row['Карта'],
            }
        )

        # Обрабатываем M2M поля с правильными lookup и code полями
        self.process_m2m(row['Возрастные ступени'], route.age_groups, AgeGroup, lookup_field='code', code_field='code')
        self.process_m2m(row['Сезоны'], route.seasons, Season, lookup_field='name')  # У Season нет code
        self.process_m2m(row['Навыки'], route.skills, Skill, lookup_field='code', code_field='code')

    def process_m2m(self, values, m2m_field, model, lookup_field='name', code_field=None):
        new = set()

        for value in str(values).split(','):
            value = value.strip()
            if not value:
                continue

            normalized = ' '.join(value.lower().split())

            obj = None
            for o in model.objects.all():
                field_value = getattr(o, lookup_field)
                if field_value:
                    normalized_field_value = ' '.join(field_value.lower().split())
                    if normalized_field_value == normalized:
                        obj = o
                        break

            if not obj:
                create_kwargs = {lookup_field: value}

                if code_field:
                    last_code_obj = model.objects.order_by(f'-{code_field}').first()
                    if last_code_obj:
                        last_code = getattr(last_code_obj, code_field)
                        if last_code and last_code.isdigit():
                            new_code = str(int(last_code) + 1)
                        else:
                            new_code = '1'
                    else:
                        new_code = '1'
                    create_kwargs[code_field] = new_code

                obj = model.objects.create(**create_kwargs)

            new.add(obj)

        m2m_field.set(new)


@admin.register(Route)
class RouteAdmin(ExportImportMixin, admin.ModelAdmin):
    list_display = ('name', 'duration', 'distance')
    change_list_template = 'admin/routes_change_list.html'


admin.site.register(AgeGroup)
admin.site.register(Season)
admin.site.register(Skill)
admin.site.register(Comment)
